from io import StringIO
import os
import re
from datetime import datetime
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from openpyxl import Workbook, load_workbook
import pandas as pd
import json
import logging_error as log


def get_list(pat):
    list_path = []
    for path, dirs, files in os.walk(pat):
        # print(path, dirs, files)
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if re.search(r"60(.*)BR(.*)", filename) and filename.endswith(".pdf"):
                list_path.append(os.path.join(path, filename))
            else:
                continue
    return list_path


def parse_pdf(file):
    output_string = StringIO()
    with open(file, 'rb') as in_file:
        parser = PDFParser(in_file)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)

    return output_string.getvalue()



# root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/60_Construction/10_Sx_Input/30_Sx_Project_Documentation/10_Mechanical_Engineering_Project/40_Piping_Iso")
root_path = os.path.normpath("D:/00_herne/test/workfiles")
#save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
#file_name = "pdf_extracted_br"
time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
search_for = "BR"
exclude = ["00_Archive", "01_Archive", "00_Document_templates"]
log_file = "log_error"
log = log.get_logger(log_file)

xls_file = "D:/00_herne/00_ISO_list_2020-10-01.xlsx"
wb = load_workbook(xls_file, read_only=False)
sh1 = wb["ISO_LIST"]   # isometric line list

line_list = next(i for i in sh1.iter_cols(min_row=1, max_row=2970, min_col=1, max_col=1))    # line kks list

data = {
    "KKS": [],
    "REV": [],
    "DATE": [],
    "FILE": [],
    "PATH": [],
    "KKS_connected": [],
    "REV_connected": [],
    "DATE_connected": [],
    "FILE_connected": [],
    "PATH_connected": []
}

for file in get_list(root_path):
    try:
        raw_pdf = parse_pdf(file)
        raw_list = list(map(str, raw_pdf.split()))
        pattern = r"60(.*){}(.*)".format(search_for)
        for s in raw_list:
            if re.search(pattern, s):
                dttm = datetime.fromtimestamp(os.path.getmtime(file)).strftime("%d.%m.%Y %H:%M")
                data["KKS"].append(os.path.basename(file).split("_")[0])
                data["REV"].append(os.path.basename(file).split(".")[0][-1])
                data["DATE"].append(dttm)
                data["FILE"].append(os.path.basename(file))
                data["PATH"].append(file)
                data["KKS_connected"].append(s.split("_")[0])
                for kks in line_list:
                    if s.split("_")[0] == kks.value:
                        data["REV_connected"].append(kks.offset(0, 1).value)
                        data["DATE_connected"].append(kks.offset(0, 2).value)
                        data["FILE_connected"].append(kks.offset(0, 3).value)
                        data["PATH_connected"].append(kks.offset(0, 4).value)
                    # else:
                    # data["REV_connected"].append("-")
                    # data["DATE_connected"].append("-")
                    # data["FILE_connected"].append("-")
                    # data["PATH_connected"].append("-")
    except Exception as err:
        log.exception(f"{file} -- {err}", exc_info=True)


with open("data_file.json", "w") as write_file:
    json.dump(data, write_file, indent=4)


# df = pd.DataFrame(data)
