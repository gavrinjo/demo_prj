from pathlib import Path
from openpyxl import load_workbook
import win32com.client as w3c
import shutil
import re
import logging_error


def export_pdf(src, dst):
    excel = w3c.Dispatch("Excel.Application")
    excel.Visible = False   # True if one wants to Excel app be opened
    excel.DisplayAlerts = False  # True if one wants to display excel alerts
    try:
        wb = excel.Workbooks.Open(src)
        sh = wb.Worksheets("04_Matrix")
        sh.Activate()
        # sh.Range("P:P").ColumnWidth = 4
        try:
            for cell in sh.Range("M6:M25"):
                # cell.WrapText = True
                # cell.Offset(1, 2).WrapText = True
                if cell.Value is None:
                    pass
                else:
                    for c in wps:
                        if cell.Value == c.value:
                            # cell.Value = c.offset(0, 1).value
                            cell.Offset(1, 2).Value = f"'{c.offset(0, 1).value}"
                            cell.Offset(1, 18).Value = c.offset(0, 2).value
                            cell.Offset(1, 19).Value = c.offset(0, 3).value
                        else:
                            pass

            try:
                wb.SaveAs(str(dst), FileFormat=57)
            except Exception as error:
                logger.exception(f"Failed to convert XLSX to PDF {src.parent} ({src.name} -> {dst.name}) // {error}",
                                 exc_info=True)

        except Exception as error:
            logger.exception(f"Failed to update XLSX file {src} // {error}", exc_info=True)
        finally:
            wb.Close(SaveChanges=True)
            excel.Quit()

    except Exception as error:
        logger.exception(f"Failed to open source file {src} // {error}", exc_info=True)


# root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\20_Sx_Working\\50_Workfiles")  # root folder containing files
# root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\20_Sx_Working\\50_Workfiles\\NDA\\60NDA20BR002")
root_path = Path("D:\\00_herne\\test\\workfiles")
log_file = "matrix_xls_auto_build"
exclude = ["00_Archive", "01_Archive", "00_Document_templates", "00_docs"]     # excluded folders
logger = logging_error.get_logger(log_file)

wb_wps = load_workbook(Path("D:/00_HERNE/test/WPS_rev_20201106/WPS_rev_20201106.xlsx"), read_only=False)
# sh_wps = wb_wps["Sheet1"]
# wps = next(i for i in sh_wps.iter_cols(min_row=1, max_row=3, min_col=1, max_col=1))
sh_br = wb_wps["Sheet1"]
br = next(i for i in sh_br.iter_cols(min_row=2, max_row=81, min_col=1, max_col=1))

wf_path = [x for x in root_path.glob(f"**") if any(y in x.parts for y in exclude) is not True if any(z in x.name for z in [k.value for k in br]) is True]
wps_list = [x for x in Path("D:\\00_HERNE\\test\\WPS_rev_20201106").glob("**/*")]

for wf in wf_path:
    # pdf_file = Path.joinpath(xl_file.parent, "04_Matrix.pdf")
    # export_pdf(xl_file, pdf_file)
    for w in wps_list:
        if wf.joinpath(w.name).exists():
            try:
                shutil.copy2(w, wf.joinpath(w.name))
            except Exception as error_1:
                logger.exception(f"Failed to copy file {w} // {error_1}", exc_info=True)

wb_wps.close()