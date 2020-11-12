import os
from datetime import datetime
from openpyxl import Workbook
import shutil

root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/60_Construction/10_Sx_Input/30_Sx_Project_Documentation/10_Mechanical_Engineering_Project/40_Piping_Iso")
# root_path = os.path.normpath("D:/00_herne/00_fabasoft")
exclude = ["00_Archive", "Derived Documents"]
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
sys_name = os.path.basename(os.path.normpath(root_path))

#wb = Workbook()
#sh = wb.active
#r = 1

"""for path, dirs, files in os.walk(root_path):
    # dirs[:] = [d for d in dirs if d == "Derived Documents"]
    if "Derived Documents" in path:
        # dttm = datetime.fromtimestamp(os.path.getmtime(os.path.join(path, filename))).strftime("%d.%m.%Y %H:%M")
        shutil.rmtree(path)
        # print(os.path.split(path)[0])
        # print(path)
        # sh.cell(r, 1, path)
        # sh.cell(r, 2, os.path.basename(path))
        # sh.cell(r, 2, filename)
        # sh.cell(r, 3).value = dttm
        # r += 1
        # list_path.append(os.path.join(path, filename))
    else:
        continue"""


subdir_list = next(os.walk(root_path))[1]
list_path = []


def get_list(pat):
    for path, dirs, files in os.walk(pat):
        # print(path, dirs, files)
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if filename.endswith(".pdf"):
                list_path.append(os.path.join(path, filename))
            else:
                continue
    return list_path


for i in subdir_list:
    wb = Workbook()
    sh = wb.active
    r = 1
    for file in get_list(os.path.join(root_path, i)):
        sh.cell(r, 1, os.path.basename(file))
        r += 1
    wb.save(os.path.join(save_path, f"{i}_{time_now}.xlsx"))
    wb.close()
    list_path = []
