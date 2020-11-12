from io import StringIO
import os
import re
from datetime import datetime
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from openpyxl import Workbook, load_workbook

import logging_error as log


def get_list(pat):
    list_path = []
    for path, dirs, files in os.walk(pat):
        # print(path, dirs, files)
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if re.search(r"60(.*)BR(.*)", filename) and filename.endswith(".pdf"):
                list_path.append(os.path.join(path, filename))
            else:
                continue
    return list_path


def parse_pdf(file):
    output_string = StringIO()
    with open(file, 'rb') as in_file:
        parser = PDFParser(in_file)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)

    return output_string.getvalue()



root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/60_Construction/10_Sx_Input/30_Sx_Project_Documentation/10_Mechanical_Engineering_Project/40_Piping_Iso")
# dir_list = []
# subdir_list = next(os.walk(root_path))[1]
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
# sys_name = os.path.basename(root_path)
file_name = "pdf_extracted_valves"
time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
#pdf_filename = 'I:/00_PROJECTS/32_IZ224_SIEMENS_Herne/60_Construction/05_Sx_Overview/01_PLAN_contracted/2.1.3_TimeSchedule_Piping_RevB.pdf'
search_for = "AA"
exclude = ["00_Archive", "01_Archive", "00_Document_templates"]
log_file = "log_error"
# list_path = []
#output_string = StringIO()
log = log.get_logger(log_file)
# for subdir in subdir_list:


xls_file = "D:/00_herne/00_ISO_list_2020-10-01.xlsx"
wb = load_workbook(xls_file, read_only=False)
sh1 = wb["ISO_LIST"]   # isometric line list

# line_list = next(i for i in sh1.iter_cols(min_row=1, max_row=2970, min_col=1, max_col=1))    # line kks list

line_list = get_list(root_path)

wb_new = Workbook()
sh_new = wb_new.active
sh_new.cell(1, 1, "KKS")
sh_new.cell(1, 2, "REV")
sh_new.cell(1, 3, "DATE")
sh_new.cell(1, 4, "FILE")
sh_new.cell(1, 5, "PATH")
sh_new.cell(1, 6, "KKS_connected")
sh_new.cell(1, 7, "REV_connected")
sh_new.cell(1, 8, "DATE_connected")
sh_new.cell(1, 9, "FILE_connected")
sh_new.cell(1, 10, "PATH_connected")
r = 2
#for dir in dir_list:
#for file in get_list(os.path.join(root_path, dir)):
for file in get_list(root_path):
    raw_pdf = None
    raw_list = []
    try:
        raw_pdf = parse_pdf(file)
        raw_list = list(map(str, raw_pdf.split()))
        pattern = r"60(.*){}(.*)".format(search_for)
        for s in raw_list:
            if re.search(pattern, s):
                dttm = datetime.fromtimestamp(os.path.getmtime(file)).strftime("%d.%m.%Y %H:%M")
                sh_new.cell(r, 1, os.path.basename(file).split("_")[0])
                sh_new.cell(r, 2, os.path.basename(file).split(".")[0][-1])
                sh_new.cell(r, 3, dttm)
                sh_new.cell(r, 4, os.path.basename(file))
                sh_new.cell(r, 5, file)
                sh_new.cell(r, 6, s.split("_")[0])
                sh_new.cell(r, 6, s)
                # for kks in line_list:
                #     if s.split("_")[0] == kks.value:
                #         sh_new.cell(r, 7, kks.offset(0, 1).value)
                #        sh_new.cell(r, 8, kks.offset(0, 2).value)
                #        sh_new.cell(r, 9, kks.offset(0, 3).value)
                #        sh_new.cell(r, 10, kks.offset(0, 4).value)
                # sh.cell(r, 3, raw_list[raw_list.index(s, + 1)])
                r += 1
    except Exception as err:
        log.exception(f"{file} -- {err}", exc_info=True)

wb_new.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))
wb_new.close()
# wb.close()
#list_path = []

