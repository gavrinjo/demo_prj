import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

import logging_error as log
import pdf_parser

p = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\10_Sx_Input\\30_Sx_Project_Documentation\\10_Mechanical_Engineering_Project\\40_Piping_Iso")
exclude_dir = ["00_Archive", "01_Archive", "00_Document_templates"]
search_for = "60*BR*"
search_ext = "pdf"


def dir_list(src, src_for=None, ext=None, exclude=None):

    if src_for is None:
        src_for = "*"
    if ext is None:
        ext = "*"

    f_list = list()

    for obj in src.glob(f"**/{src_for}.{ext}"):
        check_excluded = any(item in obj.parts for item in exclude)
        if check_excluded is not True:
            f_list.append(obj)

    return f_list


ff = dir_list(p, search_for, search_ext, exclude_dir)
r = 1
for i in ff:
    raw_pdf_data = pdf_parser.parse(i)
    raw_str_list = list(map(str, raw_pdf_data.split()))
    pattern = r"^((?!.*({}|{}).*).)*$".format("AA", "CP")
    for string in raw_str_list:
        if re.search(pattern, string):
            sh_new.cell(r, 1, i.stem.split("_")[0])
            sh_new.cell(r, 2, i.stem.split(".")[-1])
            sh_new.cell(r, 3, datetime.fromtimestamp(os.path.getmtime(i)).strftime("%d.%m.%Y %H:%M"))
            sh_new.cell(r, 4, i.name)
            sh_new.cell(r, 5, i)
            # sh_new.cell(r, 6, s.split("_")[0])
            # sh_new.cell(r, 6, s)
            # for kks in line_list:
            #     if s.split("_")[0] == kks.value:
            #         sh_new.cell(r, 7, kks.offset(0, 1).value)
            #        sh_new.cell(r, 8, kks.offset(0, 2).value)
            #        sh_new.cell(r, 9, kks.offset(0, 3).value)
            #        sh_new.cell(r, 10, kks.offset(0, 4).value)
            # sh.cell(r, 3, raw_list[raw_list.index(s, + 1)])
            r += 1

