
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import logging_error as log

log_file = "log_error"
log = log.get_logger(log_file)

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")

xl_file = "D:\\00_HERNE\\test\\raw_valve_list_20201116.xlsx"
xl_wb = load_workbook(xl_file, read_only=False)
xl_ws = xl_wb["Sheet1"]   # isometric line list

valve_list = next(i for i in xl_ws.iter_cols(min_row=2, max_row=275, min_col=1, max_col=1))    # valve kks list

wb_save_path = os.path.normpath("D:\\00_HERNE\\test\\")
wb_file_name = "_valves_xl_rebuild"
wb = Workbook()
ws = wb.active

c = 2
r = 1

try:
    for kks in valve_list:
        for cell in next(i for i in xl_ws[f'E{c}':f'HH{c}']):
            if cell.col_idx == 5 and cell.value is None:
                ws.cell(r, 1, kks.value)
                ws.cell(r, 2, kks.offset(0, 1).value)
                ws.cell(r, 3, kks.offset(0, 2).value)
                ws.cell(r, 4, kks.offset(0, 3).value.split(" | ")[0])
                ws.cell(r, 5, "fale kksovi")
            elif cell.col_idx >= 5 and cell.value is not None:
                ws.cell(r, 1, kks.value)
                ws.cell(r, 2, kks.offset(0, 1).value)
                ws.cell(r, 3, kks.offset(0, 2).value)
                ws.cell(r, 4, kks.offset(0, 3).value.split(" | ")[0])
                ws.cell(r, 5, cell.value)
            else:
                r -= 1
            r += 1
        c += 1

except Exception as err:
    log.exception(f"{err}", exc_info=True)

wb.save(os.path.join(wb_save_path, f"{wb_file_name}_{time_now}.xlsx"))
wb.close()
xl_wb.close()
