
# TODO izlistanje ventila it 00_Archive foldera

import dir_list_r01 as dl
import logging_error as log
import pdf_parser
import _pdf_rot as pr

import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from pathlib import Path

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")       # date/time in format as (Y-m-d_HMS)
root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\10_Sx_Input\\30_Sx_Project_Documentation\\10_Mechanical_Engineering_Project\\50_H&S_drawings")     # main path
exclude_dir = ["00_Archive", "00_archive", "01_Archive", "01_archive", "00_Document_templates", "SKID", "01_Deleted lines", "02_Red_Corex", "03_Additional_Workfiles", "Deleted"]     # excluded folders (these are skipped)

dlist = dl.dir_list(root_path, obj_type="f", src_for="60*BQ*", ext="pdf", exclude=exclude_dir)    # list of required files

wb_save_path = Path("D:\\00_HERNE\\_tracking\\")     # workbook save path
wb_file_name = "pipe_on_support_list"     # workbook save filename
wb_rev = "01"     # workbook revision
wb = Workbook()     # workbook
ws = wb.active      # workbook sheet activate
r = 1       # initial row number

log_file = "error_logfile"
log = log.get_logger(f"{wb_file_name}_{log_file}")

for file in dlist:
    try:
        raw_pdf_data = pdf_parser.parse(file)
        raw_str_list = list(map(str, raw_pdf_data.split()))
        pattern = r"(\d\d)(BR)"
        for string in raw_str_list:
            if re.search(pattern, string):
                ws.cell(r, 1, f'=HYPERLINK("{file}","Open")')
                ws.cell(r, 2, file.parts[-2])                   # system
                ws.cell(r, 3, file.stem.split("_")[0])          # KKS
                ws.cell(r, 4, file.stem.split("_")[1])          # UNID
                ws.cell(r, 5, file.stem.split("_")[-1][-1])     # revision
                ws.cell(r, 6, datetime.fromtimestamp(os.path.getmtime(file)).strftime("%d.%m.%Y %H:%M"))
                ws.cell(r, 7, file.name)
                ws.cell(r, 8, str(file))
                ws.cell(r, 9, string)
            else:
                r -= 1
            r += 1
    except Exception as error:
        log.exception(f"{file} --> {error}")
# ws.column_dimensions[get_column_letter(8)].hidden = True
wb.save(Path.joinpath(wb_save_path, f"{wb_file_name}_rev{wb_rev}_{time_now}.xlsx"))
wb.close()
