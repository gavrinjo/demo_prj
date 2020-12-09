import re
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
import dir_list_r01 as dl

root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\20_Sx_Working\\50_Workfiles")
exclude = ["00_Archive", "01_Archive", "00_Document_templates", "01_Deleted lines", "SKID"]
fn_list = list()
pattern = r"(\d\d)(BR)"

for i in dl.dir_list(root_path, obj_type="d", exclude=exclude):
    if re.search(pattern, i.name):
        fn_list.append(i)


def buy(**kwargs):
    time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    # wb_save_path = Path("D:\\00_HERNE\\_tracking")
    # wb_save_name = "_wf_list"
    wb = Workbook()  # workbook
    ws = wb.active  # workbook sheet activate
    r = 1  # initial row number

    for name, value in kwargs.items():
        ws.cell(r, column_index_from_string(value[0]), {value[1]})  # system


l1 = {'icecream': 5, 'apple': 1}
l2 = {"test": "parts[-3]"}

c1 = {
    "filename": ["A", f"{file.parts[-3]}"],
    "extension": ["B", f"parts[-3]"],
    "parent": ["C", f"parts[-3]"],
    "name": ["D", f"parts[-3]"],
    "UNID": ["E", f"parts[-3]"],
    "revision": ["F", f"parts[-3]"]
      }

buy(**c1)
