import os
import re
from datetime import datetime
from openpyxl import Workbook
# import pdfmine

# root_path = os.path.normpath("D:/00_herne/_backup/50_Workfiles")
root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/60_Construction/10_Sx_Input/30_Sx_Project_Documentation/10_Mechanical_Engineering_Project/40_Piping_Iso")
# root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/70_Welding/20_SQ_WPQR_WPS_HTP/01_SQ_WPS/Approved WPS")
# root_path = os.path.normpath("D:/00_herne/transmitali/T2020_09_25")
# dir_list = ["T2020_09_10", "T2020_09_11", "T2020_09_14"]
exclude = ["00_Archive","01_Archive", "00_Document_templates"]
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
file_name = "new_BR_pdf_list"
# sys_name = os.path.basename(os.path.normpath(root_path))
search_for = ["BR"]
searach_pattern = r"60(.*){}(.*)"

wb = Workbook()
sh = wb.active
r = 1
#for dir in dir_list:
    # for path, dirs, files in os.walk(os.path.join(root_path, dir)):
for path, dirs, files in os.walk(root_path):
    dirs[:] = [d for d in dirs if d not in exclude]
    for s in search_for:
        for filename in files:
            if re.search(searach_pattern.format(s), filename) and filename.endswith(".pdf"):
                dttm = datetime.fromtimestamp(os.path.getmtime(os.path.join(path, filename))).strftime("%d.%m.%Y %H:%M")
                sh.cell(r, 1, filename.split("_")[0])
                sh.cell(r, 2, filename.split(".")[0][-1])
                sh.cell(r, 3, dttm)
                sh.cell(r, 4, filename)
                sh.cell(r, 5, path)
                # sh.cell(r, 1, "ren")
                # sh.cell(r, 3, '"' + os.path.join(path, filename) + '"')
                # sh.cell(r, 5, '"' + f"05_{filename}" + '"')
                # sh.cell(r, 1, "echo y | del")
                # sh.cell(r, 2, os.path.split(os.path.split(path)[0])[1])
                # sh.cell(r, 2, os.path.basename(path))
                # sh.cell(r, 2, os.path.join(path, filename))
                # sh.cell(r, 3).value = dttm
                # os.remove(os.path.join(path, filename))
                r += 1
            else:
                continue
wb.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))

