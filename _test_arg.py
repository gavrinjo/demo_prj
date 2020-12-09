import re
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
import dir_list_r01 as dl

root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\20_Sx_Working\\50_Workfiles")
exclude = ["00_Archive", "01_Archive", "00_Document_templates", "01_Deleted lines", "SKID"]
fn_list = list()
pattern = r"(\d\d)(BR)"
"""for i in dl.dir_list(root_path, obj_type="d", exclude=exclude):
    if re.search(pattern, i.name):
        fn_list.append(i)

a1 = ws.cell(r, 1, file.parts[-3])  # system
"""


def test_a(func):

    def wrapper(save_path, save_name, **kwargs):

        for file in fn_list:
            for k in kwargs:
                func()

    return wrapper

a1 = {"collumn": "A", "obj": "parts[-3]"}


def proba(save_path, save_name, **kwargs):
    time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    # wb_save_path = Path("D:\\00_HERNE\\_tracking")
    # wb_save_name = "_wf_list"
    wb = Workbook()  # workbook
    ws = wb.active  # workbook sheet activate
    r = 1  # initial row number
    for file in fn_list:
        for i in :
            ws.cell(r, column_index_from_string(i), file.parts[-3])  # system
            ws.cell(r, 2, file.parts[-2])  # pipe
            ws.cell(r, 3, file.stem.split("_")[0])  # support point
            # ws.cell(r, 4, file.stem.split("_")[1])          # support point unid
            # ws.cell(r, 5, file.stem.split("_")[-1][-1])     # revision
            # ws.cell(r, 6, datetime.fromtimestamp(os.path.getmtime(str(file))).strftime("%d.%m.%Y %H:%M"))
            # ws.cell(r, 7, file.name)                        # filename
            ws.cell(r, 8, str(file))  # file path
            r += 1
    wb.save(Path.joinpath(save_path, f"{save_name}_{time_now}.xlsx"))
    wb.close()







# TODO simple decorator (for testing)


def my_deco(func):

    def wrapper():

        print("start with")
        func()
        print("end with")

    return wrapper


@my_deco
def test_deco():
    print("test decorator")

# test_deco = my_deco(test_deco)
# test_deco()


