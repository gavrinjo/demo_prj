import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from win32com import client
import logging_error
from itertools import groupby, cycle


def export_pdf(src, dst, sheet=None):
    exclude_sheets = ["Weldmap", "Linelist", "Sheet4"]
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False   # True if one wants to Excel app be opened
    excel.DisplayAlerts = False  # True if one wants to display excel alerts
    wb = excel.Workbooks.Open(src)
    wb_sheet_list = [sheet.Name for sheet in wb.Sheets if sheet.Name not in exclude_sheets]
    # ws_index_list = [1]
    for i in wb_sheet_list:
        if sheet is None:
            wb.WorkSheets(wb_sheet_list.index(i) + 1).Select()
            wb.ActiveSheet.ExportAsFixedFormat(0, dst)
        elif sheet == "04_Matrix":
            wb.WorkSheets(wb_sheet_list.index(sheet) + 1).Select()
            wb.ActiveSheet.ExportAsFixedFormat(0, dst)
    wb.Close()
    excel.Quit()


def vt_temp(src):
    l = []
    wb_local = load_workbook(src)
    sh4_local = wb_local["04_Matrix"]  # MATRIX Template

    for cell_obj in sh4_local["Y6":"Y27"]:
        for cell in cell_obj:
            if cell.value == "_" and cell.offset(0, -23).value == cell.offset(1, -23).value or None:
                cell.value = float(0.05)
                l.append(int(cell.coordinate.split("Y")[1]))
    try:
        temp_list = cycle(l)
        next(temp_list)
        groups = groupby(l, key=lambda j: j + 1 == next(temp_list))
        for k, v in groups:
            if k:
                group = list(tuple(v) + (next((next(groups)[1])),))
                sh4_local.merge_cells(f"Y{group[0]}:Y{group[-1]}")
    except Exception as err:
        logger.exception(f"{xls_file} -- {err}", exc_info=False)
    wb_local.save(src)
    wb_local.close()


# xls_file = "D:/00_herne/00_template/WF_temp.xlsx"  # .xlsx template file with data
xls_file = os.path.normpath("D:/00_herne/test/workfiles/EGC/60EGC10BR050/60EGC10BR050.xlsx")
# root_path = "J:/32_IZ224_SIEMENS_Herne/60_Construction/20_Sx_Working/50_Workfiles"  # root folder containing files
root_path = "D:/00_herne/test/workfiles"
log_file = "_matrix_log_file"
exclude = ["00_Archive", "01_Archive", "00_Document_templates"]     # excluded folders
# search_for = ["BR"]
# search_pattern = r"(.*)60(.*){}(.*)"
"""
wb = load_workbook(xls_file, read_only=False)
sh1 = wb["06_1_VT"]   # VT Template
sh2 = wb["Weldmap"]   # Weldmap
sh3 = wb["Linelist"]   # Linelist
sh4 = wb["04_Matrix"]   # MATRIX Template

# ref_vt_list = path_list(root_path)    # list of kks with VT file
weldmap = next(i for i in sh2.iter_cols(min_row=2, max_row=7305, min_col=1, max_col=1))     # weldmap kks list
linelist = next(i for i in sh3.iter_cols(min_row=2, max_row=2952, min_col=1, max_col=1))    # line kks list
"""
logger = logging_error.get_logger(log_file)

"""
for kks in ref_vt_list:
    try:
        vt_temp(kks)
    except Exception as error:
        logger.exception(f"{kks} -- {error}", exc_info=False)
    else:
        wb_dst = os.path.join(kks, f"{os.path.basename(kks)}.xlsx")
        replace_sheet(xls_file, wb_dst)
        pdf_file = os.path.join(kks, f"06_1_VT.pdf")
        export_pdf(xls_file, pdf_file)


del ref_vt_list, weldmap, linelist
"""
try:
    vt_temp(xls_file)
    pdf_file = os.path.join(os.path.dirname(xls_file), f"04_Matrix.pdf")
    export_pdf(xls_file, pdf_file, "04_Matrix")
except Exception as error:
    logger.exception(f"{xls_file} -- {error}", exc_info=True)
