# TODO izlistanje ventila it 00_Archive foldera

import dir_list_r01 as dl
import logging_error as log
import pdf_parser

import os
import re
import shutil as sh
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

log_file = "error_logfile"
log = log.get_logger(log_file)

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")       # date/time in format as (Y-m-d_HMS)
xl_file = Path("D:\\00_HERNE\\test\\valve_copy_list.xlsx")     # main path

wb = load_workbook(xl_file, read_only=False)     # workbook
ws = wb["Sheet1"]     # workbook sheet activate
r = 1       # initial row number

src_list = next(i for i in ws.iter_cols(min_row=1, max_row=1326, min_col=1, max_col=1))    # source valve files list
for valve in src_list:
    try:
        sh.copy2(valve.value, valve.offset(0, 2).value)
    except Exception as error:
        log.exception(f"{valve.value} --> {error}", exc_info=True)

wb.close()
