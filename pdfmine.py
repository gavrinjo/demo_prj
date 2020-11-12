import PyPDF2
import textract
import re
import os
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from openpyxl import Workbook
from datetime import datetime
from colorama import init, Fore
import pandas as pd
init()

def searchInPDF(filename, key):
    pdfFileObj = open(filename, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    num_pages = pdfReader.numPages
    count = 0
    text = ""
    while count < num_pages:
        pageObj = pdfReader.getPage(count)
        count += 1
        text += pageObj.extractText()
        # text += pageObj.getObject()
    if text != "":
        text = text
    else:
        text = textract.process(filename, method='tesseract', language='eng')
    tokens = word_tokenize(text)
    punctuation = ['(', ')', ';', ':', '[', ']', ',']
    stop_words = stopwords.words('english')
    keywords = [word for word in tokens if not word in stop_words and not word in punctuation]
    return keywords


def get_list(pat):
    for path, dirs, files in os.walk(pat):
        # print(path, dirs, files)
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if filename.endswith(".pdf"):
                list_path.append(os.path.join(path, filename))
            else:
                continue
    print(list_path)
    return list_path


#try:
#    os.remove("temp.xlsx")
#except FileNotFoundError as err:
#    print(Fore.RED, "\nThere is no 'temp.xlsx' file to delete.\nError message -->", err, Fore.RESET)

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
root_path = os.path.normpath("J:/32_IZ224_SIEMENS_Herne/60_Construction/10_Sx_Input/30_Sx_Project_Documentation/10_Mechanical_Engineering_Project/40_Piping_Iso/MBV")
save_path = os.path.normpath("D:/00_herne/01_py_script_export")
sys_name = os.path.basename(os.path.normpath(root_path))
# pdf_filename = 'D:/01_test/40_Piping_Iso/NDA/60NDA20BR903/60NDA20BR903_518159962_RevA.pdf'
search_for = ['BR']
exclude_words = []
exclude = ["00_Archive", "01_Archive"]
list_path = []
# search_for.__contains__()
wb = Workbook()
sh = wb.active
r = 1
for z in get_list(root_path):
    try:
        for h in search_for:
            for k in searchInPDF(z, h):
                if re.search(h + ".+", k):
                #if search_for in k and not search_for in exclude_words:
                    polist = [m.start() for m in re.finditer(h, k)]
                    for i in polist:
                        sh.cell(r, 1, os.path.basename(z))
                        # sh.cell(r, 2, k[i:i + 16])
                        sh.cell(r, 2, k[i - 7:i] + k[i:i + 33])
                        r += 1
    except PyPDF2.utils.PdfReadError as err:
        print(Fore.RED, "\nCorrupted PDF file >>>> " + z,"\nError message -->", err, Fore.RESET)
wb.save(os.path.join(save_path, f"{sys_name}_{time_now}.xlsx"))

#xdata = pd.read_excel("temp.xlsx", header=None)
#xdata.drop_duplicates(inplace=True)
#xdata.to_excel(os.path.join(save_path, f"{sys_name}_{time_now}.xlsx"), index=False, header=False)

#os.remove("temp.xlsx")

