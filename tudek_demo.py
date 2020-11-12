import os
import os.path as pt
import re
import itertools
from glob import glob
from datetime import datetime
from openpyxl import Workbook

import logging_error as log

root_path = "D:/00_herne/test/workfiles"
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
file_name = "tudek_test_list"
log_file = "log_error"
ext = ["pdf", "xlsx"]
log = log.get_logger(log_file)


def file_list(pat, extension=None, exclude=None):
    list_path = []
    for path, dirs, files in os.walk(pat):
        if exclude is not None:
            dirs[:] = [d for d in dirs if d not in exclude]
        else:
            pass
        if extension is not None:
            for filename in files:
                for e in extension:
                    if filename.endswith(e):
                        list_path.append(os.path.normpath(os.path.join(path, filename)))
                    else:
                        continue
        else:
            for filename in files:
                list_path.append(os.path.normpath(os.path.join(path, filename)))
    return list_path



def search_files(path):

    wb = Workbook()
    sh = wb.active
    r = 1
    offset = 0
    l = glob(f"{path}/**/*.{s if not extension else '*'}", recursive=True)
    for group in [list(g) for _, g in itertools.groupby(l, lambda x: os.path.dirname(x))]:
        for name in group:
            sh.cell(r, 1, os.path.dirname(name))
            if name.endswith(extension[0]) or name.endswith(extension[0].upper()):
                sh.cell(r, 2, os.path.basename(name))
            elif name.endswith(extension[1]) or name.endswith(extension[1].upper()):
                sh.cell(r, 3, os.path.basename(name))
            elif name.endswith(extension[2]) or name.endswith(extension[2].upper()):
                sh.cell(r, 4, os.path.basename(name))
            else:
                sh.cell(r, 5, os.path.basename(name))
            r += 1

    wb.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))


"""try:
    search_files(root_path, ext)
except Exception as err:
    log.exception(f"{root_path} -- {err}", exc_info=False)"""


k = file_list(root_path, extension=["xlsx"], exclude=["60EGC10BR001"])

