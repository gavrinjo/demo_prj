from io import StringIO
import os
import re
from datetime import datetime
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from openpyxl import Workbook

import logging_error as log


def parse_pdf(file):
    output_string = StringIO()
    with open(file, 'rb') as in_file:
        parser = PDFParser(in_file)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)

    return output_string.getvalue()


pdf_file = os.path.normpath("I:/00_PROJECTS/32_IZ224_SIEMENS_Herne/60_Construction/05_Sx_Overview/02_Termin_plan_bulk_piping_erection/Time Schedule/2.1.3_TimeSchedule_Piping_RevB.pdf")

save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
file_name = "pdf_extracted_br"
time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
log_file = "log_error"

log = log.get_logger(log_file)

# wb = Workbook()
# sh = wb.active
# r = 1

raw_list = []
try:
    raw_pdf = parse_pdf(pdf_file)
    raw_list = list(map(str, raw_pdf.split()))
    # for s in raw_list:
        # sh.cell(r, 1, os.path.basename(file).split(".")[0])
        # sh.cell(r, 3, os.path.basename(file))
        # sh.cell(r, 2, s)
        # sh.cell(r, 3, raw_list[raw_list.index(s, + 1)])
        #sh.cell(r, 1, )
        #r += 1
    print(raw_list)
except Exception as err:
    log.exception(f"{err}", exc_info=True)

# wb.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))
# wb.close()
#list_path = []
