
import shutil as sh
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from pathlib import Path

import logging_error as log

log_file = "error_logfile"
log = log.get_logger(log_file)

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")       # date/time in format as (Y-m-d_HMS)
# xl_file = Path("D:\\00_HERNE\\test\\test_copy.xlsx")     # xlsx file


def copy_xl(src, col1, col2, col3):

    """
    Copy files by list in excel document.

    Parameters:
        src (Pathlib object): The pathlib object of source xlsx file.
        col1 (str): Column containing source path + filename string.
        col2 (str): Column containing destination path + filename string.
        col3 (str): Column containing log data were copy was successful or not.

    """

    wb = load_workbook(src, read_only=False)  # workbook
    ws = wb["Sheet1"]  # workbook sheet activate

    src_column = column_index_from_string(col1)
    dst_column = column_index_from_string(col2)
    log_column = column_index_from_string(col3)

    for row in range(1, ws.max_row + 1):
        try:
            if ws.cell(row, src_column).value is not None:
                if not Path(ws.cell(row, dst_column).value).parent.exists():
                    Path(ws.cell(row, dst_column).value).parent.mkdir(parents=True, exist_ok=True)
                sh.copy2(ws.cell(row, src_column).value, ws.cell(row, dst_column).value)
            else:
                pass
        except Exception as error:
            log.exception(f"error copy {get_column_letter(src_column)}{row} --> {get_column_letter(dst_column)}{row} {error}", exc_info=True)
            ws.cell(row, log_column, f"error copy {get_column_letter(src_column)}{row} --> {get_column_letter(dst_column)}{row} {error}")
        else:
            ws.cell(row, log_column, "successful copy")

    wb.save(src)
    wb.close()
