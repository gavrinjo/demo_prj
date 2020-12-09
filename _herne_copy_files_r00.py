
# TODO copy KKS (BR(branch), BQ(support), AA(valve)) from source input folder to WF(work file) folders

import logging_error as log
import pdf_parser

import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path


log_file = "error_logfile"
log = log.get_logger(log_file)


def test_g(src, pattern, save_path, save_name):

    """
    Copy files by list in excel document.

    Parameters:
        src (Path lib object): The path lib object of root source directory or source pdf file.
        pattern (str): Regular expression search pattern ex. r"..." .
        save_path (Path lib object): Workbook save directory as path lib object.
        save_name (str): Workbook save filename.

    """
    source = list()

    if type(src) is not list:
        source.append(src)
    else:
        source = src

    time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")  # date/time in format as (Y-m-d_HMS)
    wb = Workbook()  # workbook
    ws = wb.active  # workbook sheet activate
    r = 1  # initial row number

    for file in source:
        try:
            raw_pdf_data = pdf_parser.parse(file)
            raw_str_list = list(map(str, raw_pdf_data.split()))
            for string in raw_str_list:
                if re.search(pattern, string):
                    ws.cell(r, 1, file.stem.split("_")[0])
                    ws.cell(r, 2, file.stem.split("_")[1])
                    ws.cell(r, 3, file.stem.split("_")[2][-1])
                    ws.cell(r, 4, datetime.fromtimestamp(os.path.getmtime(file)).strftime("%d.%m.%Y %H:%M"))
                    ws.cell(r, 5, file.name)
                    ws.cell(r, 6, str(file))
                    ws.cell(r, 7, string)
                else:
                    r -= 1
                r += 1
        except Exception as error:
            log.exception(f"{file} --> {error}", exc_info=True)

    wb.save(Path.joinpath(save_path, f"{save_name}_{time_now}.xlsx"))
    wb.close()
