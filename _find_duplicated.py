import sys
import os
import hashlib
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
import dir_list_r01 as dl
import logging_error as log


def chunk_reader(fobj, chunk_size=1024):
    """Generator that reads a file in chunks of bytes"""
    while True:
        chunk = fobj.read(chunk_size)
        if not chunk:
            return
        yield chunk


def check_for_duplicates(paths, hash=hashlib.sha1, error_log=log):
    hashes = {}
    time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    exclude = ["00_Archive", "00_archive", "01_Archive", "01_archive", "00_Document_templates", "01_Deleted lines", "02_Red_Corex",
               "03_Additional_Workfiles", "SKID", "Bare"]

    wb_save_path = Path("N:\\DGAVRIC\\_ITTER")
    wb_save_name = "duplicated_files_proba"
    wb_rev = "02"  # workbook revision
    wb = Workbook()  # workbook
    ws = wb.active  # workbook sheet activate
    r = 1  # initial row number

    log_file = "error_logfile"
    logger = error_log.get_logger(f"{wb_save_name}_{log_file}")

    for path in paths:
        fn_list = dl.dir_list(path, typ="f", lookup="*", extension="*", exclude=exclude)
        for file in fn_list:
            hashobj = hash()
            for chunk in chunk_reader(open(file, 'rb')):
                hashobj.update(chunk)
            file_id = (hashobj.digest(), os.path.getsize(file))
            duplicate = hashes.get(file_id, None)
            if duplicate:
                # print("Duplicate found: %s and %s" % (file, duplicate))
                try:
                    ws.cell(r, 1, f'=HYPERLINK("{file}","Open")')
                    ws.cell(r, 2, file.name)
                    ws.cell(r, 3, str(file))  # file path
                    ws.cell(r, 5, f'=HYPERLINK("{duplicate}","Open")')
                    ws.cell(r, 6, duplicate.name)
                    ws.cell(r, 7, str(duplicate))  # file path
                    r += 1
                except Exception as error:
                    logger.exception(f"{wb_save_name}_{file} --> {error}")
            else:
                hashes[file_id] = file

    wb.save(Path.joinpath(wb_save_path, f"{wb_rev}_{wb_save_name}_{time_now}.xlsx"))
    wb.close()


'''
if sys.argv[1:]:
    check_for_duplicates(sys.argv[1:])
else:
    print("Please pass the paths to check as parameters to the script")'''

rp = ["N:\\DGAVRIC\\_ITTER\\08_Ax_Tender_Documentation"]
check_for_duplicates(rp)
