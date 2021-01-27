import re
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
import dir_list_r01 as dl
import logging_error as log

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\20_Sx_Working\\50_Workfiles")
# root_path = Path("J:\\32_IZ224_SIEMENS_Herne\\60_Construction\\10_Sx_Input\\30_Sx_Project_Documentation\\10_Mechanical_Engineering_Project\\50_H&S_drawings")
exclude = ["00_Archive", "00_archive", "01_Archive", "01_archive", "00_Document_templates", "02_Red_Corex", "03_Additional_Workfiles", "SKID", "Deleted"]

# ls = dl.dir_list(root_path, ext="pdf", exclude=exclude)     # list of valves
"""
# fn_list = list()
pattern = r"(\d\d)(BQ)"
for i in dl.dir_list(root_path, exclude=exclude):
    if re.search(pattern, i.name):
        fn_list.append(i)
"""

fn_list = dl.dir_list(root_path, obj_type="f", src_for="60*BQ*", ext="pdf", exclude=exclude)
wb_save_path = Path("D:\\00_HERNE\\_tracking\\script_exports")
wb_save_name = "support_files_in_wf_list_with_deleted_lines"
wb_rev = "05"     # workbook revision
wb = Workbook()     # workbook
ws = wb.active      # workbook sheet activate
r = 1       # initial row number
# patt = r"(\d{9})"

log_file = "error_logfile"
log = log.get_logger(f"{wb_save_name}_{log_file}")

for file in fn_list:
    if "redmark" not in str(file):
        try:
            ws.cell(r, 1, f'=HYPERLINK("{file}","Open")')
            ws.cell(r, 2, file.parts[-3])                   # system
            ws.cell(r, 3, file.parts[-2])                   # kks
            ws.cell(r, 4, file.stem.split("_")[0])          # support point
            ws.cell(r, 5, file.stem.split("_")[1])          # support point unid
            try:
                ws.cell(r, 6, file.stem.split("_")[2][-1])  # revision
            except IndexError:
                ws.cell(r, 6, file.stem.split("-")[1][-1])  # revision
            ws.cell(r, 7, datetime.fromtimestamp(os.path.getmtime(str(file))).strftime("%d.%m.%Y %H:%M"))
            ws.cell(r, 8, file.name)                        # filename
            ws.cell(r, 9, str(file))                        # file path
            r += 1
        except Exception as error:
            log.exception(f"{wb_save_name}_{file} --> {error}")
    else:
        pass

wb.save(Path.joinpath(wb_save_path, f"{wb_save_name}_rev{wb_rev}_{time_now}.xlsx"))
wb.close()
