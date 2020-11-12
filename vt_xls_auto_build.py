import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from win32com import client
import logging_error


def get_vt_list(pat):
    search_pattern = r"06_1_VT"
    file_ext = ".pdf"
    list_path = []
    for path, dirs, files in os.walk(pat):
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if re.search(search_pattern, filename) and filename.endswith(file_ext):
                list_path.append(os.path.normpath(path))
            else:
                continue
    return list_path


def export_pdf(src, dst):
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = False   # True if one wants to Excel app be opened
    excel.DisplayAlerts = False  # True if one wants to display excel alerts
    wb = excel.Workbooks.Open(src)
    ws_index_list = [1]
    wb.WorkSheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, dst)
    wb.Close()
    excel.Quit()


def replace_sheet(src, dst):
    excel = client.Dispatch("Excel.Application")
    excel.Visible = False   # True if one wants to Excel app be opened
    excel.DisplayAlerts = False     # True if one wants to display excel alerts
    wb1 = excel.Workbooks.Open(src)
    ws1 = wb1.Worksheets(1)

    if os.path.isfile(dst):
        wb2 = excel.Workbooks.Open(dst)
        wb2_sheet_list = [sheet.Name for sheet in wb2.Sheets]
        if "06_1_VT" in wb2_sheet_list:
            wb2.Worksheets(wb2_sheet_list.index("06_1_VT") + 1).Delete()

        if "07_Cleanlines" in wb2_sheet_list:
            ws1.Copy(Before=wb2.Worksheets(wb2_sheet_list.index("07_Cleanlines")))
        else:
            ws1.Copy(Before=wb2.Worksheets(1))
    else:
        wb2 = Workbook()
        wb2.save(dst)
        wb2.close()
        wb2 = excel.Workbooks.Open(dst)
        ws1.Copy(Before=wb2.Worksheets(1))

    wb2.Close(SaveChanges=True)
    excel.Quit()


def vt_temp(src):
    r = 0
    wp = []
    for cell_obj in sh1["A30":"A50"]:
        for cell in cell_obj:
            cell.value = None
    for wn in weldmap_kks:
        if os.path.basename(src) == wn.value:
            sh1["A{}".format(30 + r)] = wn.offset(0, 1).value
            r += 1
            wp.append(wn.offset(0, 16).value)
            wp_set = set(wp)
            sh1["K19"] = "; ".join(map(str, wp_set))

    for ll_kks in linelist_kks:
        if os.path.basename(src) == ll_kks.value:
            sh1["I13"] = ll_kks.offset(0, 4).value
            sh1["K15"] = os.path.basename(src)
            sh1["K17"] = ll_kks.offset(0, 3).value

    wb.save(xls_file)


xls_file = "D:/00_herne/00_template/VT.xlsx"  # .xlsx file with data
root_path = "J:/32_IZ224_SIEMENS_Herne/60_Construction/20_Sx_Working/50_Workfiles"  # root folder containing files
# root_path = "D:/00_herne/test/workfiles"
log_file = "VT_bulk_print"
log_path = "D:/00_herne/01_py_script_export/logs"
exclude = ["00_Archive", "01_Archive", "00_Document_templates"]     # excluded folders

wb = load_workbook(xls_file, read_only=False)
sh1 = wb.worksheets[0]   # Template
sh2 = wb.worksheets[1]   # Weldmap
sh3 = wb.worksheets[2]   # Linelist

ref_vt_list = get_vt_list(root_path)    # list of kks with VT file
weldmap_kks = next(i for i in sh2.iter_cols(min_row=2, max_row=7305, min_col=1, max_col=1))     # weldmap kks list
linelist_kks = next(i for i in sh3.iter_cols(min_row=2, max_row=2952, min_col=1, max_col=1))    # line kks list

logger = logging_error.get_logger(log_file)

for kks in ref_vt_list:
    try:
        vt_temp(kks)
        wb_dst = os.path.join(kks, f"{os.path.basename(kks)}.xlsx")
        replace_sheet(xls_file, wb_dst)
        pdf_file = os.path.join(kks, f"06_1_VT.pdf")
        export_pdf(xls_file, pdf_file)
    except Exception as error:
        logger.exception(f"{kks} -- {error}", exc_info=True)

wb.close()
del ref_vt_list, weldmap_kks, linelist_kks
