
import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime

import logging_error

"""wf_cl_df = "D:/00_herne/_export_SmartSheet/WF_CONTROL_LIST_20200923.xlsx"   # work file control list data frame
df = pd.read_excel(wf_cl_df, sheet_name="WF_CONTROL_LIST")
"""
# print(df[df["BRANCH"] == "60LBA20BR303"].values[0])

# --------CODE:--------
"""time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
# root_path = "D:/00_herne/test/workfiles"
root_path = "J:/32_IZ224_SIEMENS_Herne/60_Construction/20_Sx_Working/50_Workfiles"
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
file_name = "Matix_list"
search_pattern = r"60(.*)BR(.*)"
file_ext = ".xlsx"
exclude = ["00_Archive","01_Archive", "00_Document_templates"]
sheet_name = "04_Matrix"
log_file = "Matrix_print"
rectify_cells = ["AF29", "AG27"]
logger = logging_error.get_logger(log_file)"""


def path_list(pat):
    
    list_path = []
    for path, dirs, files in os.walk(pat):
        dirs[:] = [d for d in dirs if d not in exclude]
        for filename in files:
            if re.search(search_pattern, filename) and filename.endswith(file_ext):
                list_path.append(os.path.normpath(os.path.join(path, filename)))
            else:
                continue
    return list_path


def rectify_dates(xl_file, sheet):
    wb = load_workbook(xl_file, read_only=False)
    ws = wb[f"{sheet}"]
    global r

    writer = pd.ExcelWriter(xl_file, engine="openpyxl")
    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    branch = os.path.basename(os.path.dirname(xl_file))
    date_value = pd.Timestamp(np.datetime64(df[df["BRANCH"] == f"{branch}"].filter(["WF DATE"]).values[0][0])).to_pydatetime()
    for i in rectify_cells:
        if ws[f"{i}"].value is not None:
            sh.cell(r, 1, os.path.basename(xl_file).split(".")[0])
            sh.cell(r, 2, i)
            if ws[f"{i}"] == date_value:
                sh.cell(r, 3, "True")
            else:
                sh.cell(r, 3, date_value.strftime("%d.%m.%Y"))
            sh.cell(r, 4, xl_file)
            r += 1


    # TODO: write newer date to .xlsx file
    # df[df["Line KKS"] == "60LBA20BR004"].to_excel(writer, sheet_name="04_Matrix", startrow=5, header=False, index=False, float_format="%.2f")


"""wbook = Workbook()
sh = wbook.active
r = 1

for file in path_list(root_path):
    try:
        rectify_dates(file, sheet_name)
    except Exception as err:
        logger.exception(f"{file} -- {err}", exc_info=True)


wbook.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))
wbook.close()
"""