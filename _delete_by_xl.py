
import shutil as sh
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from pathlib import Path

import logging_error as log

log_file = "error_logfile"
log = log.get_logger(log_file)

time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")       # date/time in format as (Y-m-d_HMS)


def copy_xl(src, col1, col3):

    """
    Copy files by list in excel document.

    Parameters:
        src (Pathlib object): The pathlib object of source xlsx file.
        col1 (str): Column containing source path + filename string.
        col3 (str): Column containing log data were copy was successful or not.

    """

    wb = load_workbook(src, read_only=False)  # workbook
    ws = wb["Sheet"]  # workbook sheet activate

    src_column = column_index_from_string(col1)
    log_column = column_index_from_string(col3)

    for row in range(1, ws.max_row + 1):
        try:
            if ws.cell(row, src_column).value is not None:
                os.remove(ws.cell(row, src_column).value)
                if Path(ws.cell(row, src_column).value).exists():
                    ws.cell(row, log_column, "file successfully removed")
                else:
                    pass
            else:
                pass
        except Exception as error:
            log.exception(f"error delete file {get_column_letter(src_column)}{row} {error}", exc_info=False)
            ws.cell(row, log_column, f"error delete {get_column_letter(src_column)}{row}")

    wb.save(src)
    wb.close()


xl_file = Path("N:\\DGAVRIC\\_ITTER\\01_duplicated_files_2021-01-29.xlsx")     # xlsx file

copy_xl(xl_file, "C", "D")
