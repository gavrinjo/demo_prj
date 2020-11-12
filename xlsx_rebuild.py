import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import logging_error

file = "D:/00_herne/_export_SmartSheet/IZ224_Termin plan_Bulk piping erection_20200925.xlsx"
# save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
# time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
# file_name = "TP_line_list"

log_file = "TP_log"
log_path = "D:/00_herne/01_py_script_export/logs"

logger = logging_error.get_logger(log_file)

wb = load_workbook(file)
sh1 = wb["IZ224_Termin_plan"]   # data roster
sh1_range = next(n for n in sh1.iter_cols(min_row=1, max_row=13250, min_col=5, max_col=5))

for sh1_cell in sh1_range:
    try:
        if sh1_cell.value == "KKS":
            sh1.insert_rows(sh1_cell.row + 1)
            sh1_cell.value = sh1_cell.offset(0, -2).value
            sh1_cell.offset(1, -2).value = sh1_cell.offset(0, -2).value
            sh1_cell.offset(1, 0).value = "Erection of spools, supports and secondary steel structure (if any) and welding"
        elif sh1_cell.value == "Pipe":
            sh1_cell.value = "NDT / PWHT (if any)"
        elif sh1_cell.value == "Weld":
            sh1_cell.value = "Ready for pressure test"
        elif sh1_cell.value == "H&S":
            sh1_cell.value = "ECC"
    except Exception as err:
        logger.exception(f"{sh1_cell.coordinate} -- {err}", exc_info=True)

        # print(r)
# column_index_from_string(coordinate_from_string(sh1_cell.row)[1])
# abc = next(i for i in sh.iter_cols(min_row=2, max_row=48, min_col=1, max_col=1, values_only=True))
# print(abc)
# print(xdata)
# xdata.drop_duplicates(inplace=True)
# xdata.to_excel(os.path.join(save_path, f"{sys_name}_{time_now}.xlsx"), index=False, header=False)

# new_wb.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))
wb.save(file)
wb.close()
