import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime

file = "D:/00_herne/01_py_script_export/linelist.xlsx"
save_path = os.path.normpath("D:/00_herne/01_py_script_export/sys")
time_now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
file_name = "line_list"

wb = load_workbook(file)
sheet1 = wb["Sheet1"]   # data roster
sheet2 = wb["WA_list"]   # work area list

new_wb = Workbook()
sh = new_wb.active
r = 1

for sh2_cell in next(i for i in sheet2.iter_cols(min_row=2, max_row=48, min_col=1, max_col=1)):
    offset = 0
    for sh1_cell in next(n for n in sheet1.iter_cols(min_row=2, max_row=3191, min_col=3, max_col=3)):
        if sh2_cell.value == sh1_cell.value:
            if sh1_cell.offset(offset, 0) == sh1_cell:
                # print(sh1_cell.value)
                sh.cell(r, 1, sh1_cell.value)
                r += 1
            # print(sh1_cell.offset(0, 1).value)
            sh.cell(r, 1, sh1_cell.offset(0, 1).value)
            sh.cell(r, 2, sh1_cell.offset(0, 1).value)
            r += 1
            sh.cell(r, 1, "Erection of spools, supports and secondary steel structure (if any) and welding")
            sh.cell(r, 2, sh1_cell.offset(0, 1).value)
            r += 1
            sh.cell(r, 1, "NDT / PWHT (if any)")
            sh.cell(r, 2, sh1_cell.offset(0, 1).value)
            r += 1
            sh.cell(r, 1, "Ready for pressure test")
            sh.cell(r, 2, sh1_cell.offset(0, 1).value)
            r += 1
            sh.cell(r, 1, "ECC")
            sh.cell(r, 2, sh1_cell.offset(0, 1).value)
            offset += 1
            r += 1
            # print(r)

# abc = next(i for i in sh.iter_cols(min_row=2, max_row=48, min_col=1, max_col=1, values_only=True))
# print(abc)
# print(xdata)
# xdata.drop_duplicates(inplace=True)
# xdata.to_excel(os.path.join(save_path, f"{sys_name}_{time_now}.xlsx"), index=False, header=False)

new_wb.save(os.path.join(save_path, f"{file_name}_{time_now}.xlsx"))
new_wb.close()
